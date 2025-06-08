import datetime
import time
import threading
import os
import sys
import json
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
import pandas as pd

class ExcelReminderApp:
    def __init__(self, excel_path, time_column, content_columns=None):
        self.excel_path = excel_path
        self.time_column = time_column
        self.content_columns = content_columns or []
        self.today_data = []
        self.previous_data = []  # 存储上次数据用于对比
        self.stop_event = threading.Event()
        self.check_thread = None
        self.cache_file = 'data_cache.json'

    def load_today_data(self, print_new_records=False):
        """加载今日数据，可选打印新增记录"""
        today = datetime.date.today()
        try:
            # 尝试从缓存加载
            try:
                with open(self.cache_file, 'r') as f:
                    cache = json.load(f)
                    if cache.get('date') == str(today):
                        self.previous_data = self.today_data.copy()
                        self.today_data = cache.get('data', [])
                        if print_new_records:
                            self._print_new_records()
                        return True, f"从缓存加载 {len(self.today_data)} 条记录"
            except (FileNotFoundError, json.JSONDecodeError):
                pass

            # 从Excel加载
            if not os.path.exists(self.excel_path):
                return False, f"文件不存在: {self.excel_path}"

            file_ext = os.path.splitext(self.excel_path)[1].lower()
            if file_ext == '.xlsx':
                wb = load_workbook(self.excel_path, data_only=True)
                ws = wb.active
                time_col_idx, content_col_indices = self._get_columns(ws)
                self.previous_data = self.today_data.copy()
                self.today_data = self._parse_xlsx_rows(ws, time_col_idx, content_col_indices, today)
                wb.close()
            elif file_ext == '.xls':
                df = pd.read_excel(self.excel_path)
                self.previous_data = self.today_data.copy()
                self.today_data = self._parse_xls_data(df, today)
            else:
                return False, f"不支持的文件类型: {file_ext}"

            self._cache_data(today)
            
            # 打印新增记录
            if print_new_records:
                self._print_new_records()
                
            return True, f"成功加载 {len(self.today_data)} 条今日记录"
        except Exception as e:
            return False, f"加载失败: {str(e)}"

    def _print_new_records(self):
        """打印新增的记录"""
        if not self.previous_data:  # 首次加载
            print(f"\n首次加载 {len(self.today_data)} 条今日记录:")
            for record in self.today_data:
                self._print_record(record)
        else:  # 刷新时对比新旧数据
            old_ids = set(str(record["时间"]) + str(record.get("姓名", "")) for record in self.previous_data)
            new_records = [r for r in self.today_data if (str(r["时间"]) + str(r.get("姓名", ""))) not in old_ids]
            
            if new_records:
                print(f"\n检测到 {len(new_records)} 条新增记录:")
                for record in new_records:
                    self._print_record(record)
            else:
                print("\n没有新增记录")

    def _print_record(self, record):
        """格式化打印单条记录"""
        time_str = record["时间"].strftime("%Y-%m-%d %H:%M:%S")
        details = ', '.join([f'{k}: {v}' for k, v in record.items() if k != '时间'])
        print(f"{time_str}: {details}")

    def _get_columns(self, ws):
        time_col_idx = None
        content_col_indices = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == self.time_column:
                time_col_idx = col_idx
            if cell.value in self.content_columns:
                content_col_indices[cell.value] = col_idx
        if not time_col_idx:
            raise ValueError(f"未找到时间列: {self.time_column}")
        return time_col_idx, content_col_indices

    def _parse_xlsx_rows(self, ws, time_col_idx, content_col_indices, today):
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            time_value = row[time_col_idx - 1]
            if time_value:
                time_obj = self._parse_time(time_value)
                if time_obj.date() == today:
                    record = {'时间': time_obj}
                    for col_name, col_idx in content_col_indices.items():
                        record[col_name] = row[col_idx - 1] if col_idx - 1 < len(row) else None
                    data.append(record)
        return data

    def _parse_xls_data(self, df, today):
        if self.time_column not in df.columns:
            raise ValueError(f"未找到时间列: {self.time_column}")
        df['datetime'] = pd.to_datetime(df[self.time_column])
        today_start = datetime.datetime.combine(today, datetime.time.min)
        today_end = datetime.datetime.combine(today, datetime.time.max)
        today_df = df[(df['datetime'] >= today_start) & (df['datetime'] <= today_end)]
        data = []
        for _, row in today_df.iterrows():
            record = {'时间': row['datetime'].to_pydatetime()}
            for col in self.content_columns:
                record[col] = row[col] if col in row else None
            data.append(record)
        return data

    def _parse_time(self, time_value):
        if isinstance(time_value, str):
            try:
                return datetime.datetime.strptime(time_value, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    return datetime.datetime.strptime(time_value, '%Y-%m-%d')
                except ValueError:
                    try:
                        # 尝试其他常见格式
                        return datetime.datetime.strptime(time_value, '%m/%d/%Y %H:%M:%S')
                    except ValueError:
                        return datetime.datetime.strptime(time_value, '%m/%d/%Y')
        elif isinstance(time_value, datetime.datetime):
            return time_value
        elif isinstance(time_value, datetime.date):
            return datetime.datetime.combine(time_value, datetime.time())
        else:
            raise ValueError(f"未知时间格式: {time_value}")

    def _cache_data(self, today):
        cache = {'date': str(today), 'data': self.today_data}
        with open(self.cache_file, 'w') as f:
            json.dump(cache, f, default=str)

    def start_refreshing(self, interval=3600):
        if not self.check_thread or not self.check_thread.is_alive():
            self.check_thread = threading.Thread(target=self._refresh_loop, args=(interval,))
            self.check_thread.daemon = True
            self.check_thread.start()

    def _refresh_loop(self, interval):
        while not self.stop_event.is_set():
            success, message = self.load_today_data(print_new_records=True)
            print(f"自动刷新: {message}")
            time.sleep(interval)

    def stop_refreshing(self):
        self.stop_event.set()
        if self.check_thread and self.check_thread.is_alive():
            self.check_thread.join(timeout=1)

class ExcelReminderGUI:
    def __init__(self, root, excel_path, time_column, content_columns=None, subtitle=None):
        self.root = root
        self.root.title("小美的预约系统")
        self.root.geometry("800x600")
        self.root.configure(bg="#f0f0f0")

        self.app = ExcelReminderApp(excel_path, time_column, content_columns)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.subtitle = subtitle or datetime.datetime.now().strftime("%Y年%m月%d日")
        self.tree = None
        self.auto_refresh_var = tk.BooleanVar(value=False)
        self.status_var = tk.StringVar(value="准备加载数据...")

        self.create_widgets()
        
        # 先加载数据并打印到终端
        success, message = self.app.load_today_data(print_new_records=True)
        print(message)
        
        # 再更新GUI
        self._init_tree_if_needed()
        self._update_tree_data()
        self.status_var.set(message)

    def create_widgets(self):
        self._create_title_frame()
        self._create_button_frame()
        self._create_status_label()

    def _create_title_frame(self):
        title_frame = tk.Frame(self.root, bg="#f0f0f0")
        title_frame.pack(pady=10, fill=tk.X)

        tk.Label(title_frame, text="小美的预约系统", font=("微软雅黑", 18, "bold"), bg="#f0f0f0").pack(side=tk.LEFT, padx=20)
        tk.Label(title_frame, text=self.subtitle, font=("微软雅黑", 10), bg="#f0f0f0", fg="#666").pack(side=tk.RIGHT, padx=20)

    def _create_button_frame(self):
        button_frame = tk.Frame(self.root, bg="#f0f0f0")
        button_frame.pack(pady=5, fill=tk.X)

        tk.Button(button_frame, text="自动刷新", command=self.toggle_auto_refresh,
                  font=("微软雅黑", 10), bg="#4CAF50", fg="white", padx=15, pady=5).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="刷新数据", command=self.load_data,
                  font=("微软雅黑", 10), bg="#2196F3", fg="white", padx=15, pady=5).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="退出", command=self.on_close,
                  font=("微软雅黑", 10), bg="#f44336", fg="white", padx=15, pady=5).pack(side=tk.LEFT, padx=5)

    def _create_status_label(self):
        status_label = tk.Label(self.root, textvariable=self.status_var,
                                font=("微软雅黑", 10), bg="#f0f0f0", fg="blue")
        status_label.pack(pady=5, fill=tk.X)

    def load_data(self):
        self.status_var.set("正在刷新数据...")
        success, message = self.app.load_today_data(print_new_records=True)
        print(message)
        self._init_tree_if_needed()
        self._update_tree_data()
        self.status_var.set(message)

    def _init_tree_if_needed(self):
        if self.tree is None:
            columns = ["时间"] + self.app.content_columns
            self.tree = ttk.Treeview(self.root, columns=columns, show="headings", selectmode='browse')
            self.tree.column("时间", width=150)
            self.tree.heading("时间", text="时间")
            for col in self.app.content_columns:
                self.tree.column(col, width=150)
                self.tree.heading(col, text=col)
            scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.tree.yview)
            self.tree.configure(yscrollcommand=scrollbar.set)
            self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=10)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=10)

    def _update_tree_data(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for record in self.app.today_data:
            values = [record["时间"].strftime("%Y-%m-%d %H:%M:%S")]
            for col in self.app.content_columns:
                values.append(record.get(col, ""))
            self.tree.insert("", tk.END, values=values)

    def toggle_auto_refresh(self):
        if self.auto_refresh_var.get():
            self.app.stop_refreshing()
            self.status_var.set("自动刷新已关闭")
        else:
            self.app.start_refreshing()
            self.status_var.set("自动刷新已开启 (每1小时)")
        self.auto_refresh_var.set(not self.auto_refresh_var.get())

    def on_close(self):
        if messagebox.askyesno("确认", "确定退出吗？"):
            self.app.stop_refreshing()
            self.root.destroy()

def main():
    excel_path = "/Users/Sun/Desktop/预约/患者管理登记表.xlsx"  # 替换为实际路径
    time_column = "复诊时间"  # Excel 中的时间列名
    content_columns = ["姓名", "处置", "余留问题"]  # 需展示的列名
    subtitle = "栋哥特约版V1.0"

    # 静默模式（仅打印数据）
    if '--silent' in sys.argv:
        app = ExcelReminderApp(excel_path, time_column, content_columns)
        success, message = app.load_today_data(print_new_records=True)
        print(message)
        if success:
            app.start_refreshing()
            try:
                while True:
                    time.sleep(1)
            except KeyboardInterrupt:
                app.stop_refreshing()
                print("程序已退出")
    # GUI模式（先打印数据，再显示界面）
    else:
        print(f"正在从 {excel_path} 加载今日数据...")
        root = tk.Tk()
        ExcelReminderGUI(root, excel_path, time_column, content_columns, subtitle)
        root.mainloop()

if __name__ == "__main__":
    main()
